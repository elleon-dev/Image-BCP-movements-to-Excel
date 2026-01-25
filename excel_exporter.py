"""
Módulo para exportar transacciones bancarias a Excel.
Organiza por meses y formatea con información de cuenta.
"""

from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exportador de transacciones a Excel."""
    
    def __init__(self, account_type: str = "Cuenta", account_number: str = "", bank_name: str = "Banco"):
        """
        Inicializa el exportador.
        
        Args:
            account_type: Tipo de cuenta (ej: "Cuenta Corriente")
            account_number: Número de cuenta
            bank_name: Nombre del banco
        """
        self.account_type = account_type
        self.account_number = account_number
        self.bank_name = bank_name
    
    def sort_transactions(self, transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Ordena transacciones por fecha descendente (más reciente primero).
        
        Args:
            transactions: Lista de transacciones
            
        Returns:
            Lista ordenada de transacciones
        """
        def parse_date(transaction):
            try:
                date_str = transaction.get('date', '01/01/2000')
                return datetime.strptime(date_str, '%d/%m/%Y')
            except:
                return datetime.min
        
        return sorted(transactions, key=parse_date, reverse=True)
    
    def group_by_month(self, transactions: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Agrupa transacciones por mes.
        
        Args:
            transactions: Lista de transacciones
            
        Returns:
            Diccionario con meses como claves y listas de transacciones como valores
        """
        grouped = {}
        
        for transaction in transactions:
            month = transaction.get('month', 'Sin mes')
            if month not in grouped:
                grouped[month] = []
            grouped[month].append(transaction)
        
        # Ordenar transacciones dentro de cada mes
        for month in grouped:
            grouped[month] = self.sort_transactions(grouped[month])
        
        return grouped
    
    def create_excel(self, transactions: List[Dict[str, Any]], output_path: str = "output/movimientos_bancarios.xlsx"):
        """
        Crea archivo Excel con las transacciones.
        
        Args:
            transactions: Lista de transacciones
            output_path: Ruta del archivo de salida
        """
        # Crear directorio de salida si no existe
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Ordenar todas las transacciones
        transactions = self.sort_transactions(transactions)
        
        # Agrupar por mes
        grouped = self.group_by_month(transactions)
        
        if not grouped:
            print("⚠️  No hay transacciones para exportar")
            return
        
        # Crear workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Estilos
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Crear una hoja por mes
        for month, month_transactions in grouped.items():
            # Nombre seguro para la hoja (máximo 31 caracteres)
            sheet_name = month[:31] if len(month) <= 31 else month[:28] + "..."
            ws = wb.create_sheet(title=sheet_name)
            
            # Título con información de cuenta
            title = f"{month} - {self.account_type}"
            if self.account_number:
                title += f" N° {self.account_number}"
            title += f" - {self.bank_name}"
            
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Encabezados de columna
            headers = ['Fecha', 'Descripción', 'Tipo', 'Monto', 'Moneda']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos de transacciones
            row_num = 4
            for transaction in month_transactions:
                # Fecha
                cell = ws.cell(row=row_num, column=1)
                cell.value = transaction.get('date', '')
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                # Descripción
                cell = ws.cell(row=row_num, column=2)
                cell.value = transaction.get('name', '')
                cell.alignment = Alignment(horizontal='left')
                cell.border = border
                
                # Tipo
                cell = ws.cell(row=row_num, column=3)
                cell.value = transaction.get('type', '').capitalize()
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                # Color basado en tipo
                if transaction.get('type') == 'cargo':
                    cell.font = Font(color="C00000")  # Rojo para cargos
                else:
                    cell.font = Font(color="00B050")  # Verde para abonos
                
                # Monto
                cell = ws.cell(row=row_num, column=4)
                amount = transaction.get('amount', 0)
                # Si es cargo, mantener el signo negativo; si es abono, valor positivo
                if transaction.get('type') == 'cargo':
                    cell.value = -abs(amount)  # Negativo para cargos
                else:
                    cell.value = abs(amount)  # Positivo para abonos
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                
                # Aplicar color al monto también
                if transaction.get('type') == 'cargo':
                    cell.font = Font(color="C00000")
                else:
                    cell.font = Font(color="00B050")
                
                # Moneda
                cell = ws.cell(row=row_num, column=5)
                cell.value = transaction.get('currency', 'S/')  # Usar moneda de la transacción
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                row_num += 1
            
            # Totales
            row_num += 1
            total_cargos = sum(abs(t['amount']) for t in month_transactions if t.get('type') == 'cargo')
            total_abonos = sum(abs(t['amount']) for t in month_transactions if t.get('type') == 'abono')
            balance = total_abonos - total_cargos
            
            # Total Cargos
            ws.cell(row=row_num, column=2).value = "Total Cargos:"
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=4).value = total_cargos
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=4).font = Font(bold=True, color="C00000")
            ws.cell(row=row_num, column=5).value = 'S/'
            
            # Total Abonos
            row_num += 1
            ws.cell(row=row_num, column=2).value = "Total Abonos:"
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=4).value = total_abonos
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=4).font = Font(bold=True, color="00B050")
            ws.cell(row=row_num, column=5).value = 'S/'
            
            # Balance
            row_num += 1
            ws.cell(row=row_num, column=2).value = "Balance:"
            ws.cell(row=row_num, column=2).font = Font(bold=True)
            ws.cell(row=row_num, column=4).value = balance
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=4).font = Font(bold=True, color="1F4E78")
            ws.cell(row=row_num, column=5).value = 'S/'
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 8
        
        # Guardar archivo
        wb.save(output_path)
        print(f"\n✅ Excel generado: {output_path}")
        print(f"📑 Hojas creadas: {len(grouped)}")
        print(f"📊 Total transacciones: {len(transactions)}")
